"""한국은행 '일일 금융시장 주요지표' 엑셀 파싱 수집기 (자금흐름 은행·증권).

RSS로 최근 게시글을 찾아 각 게시글의 엑셀 첨부('잔액' 우선)를 내려받아
'2. 금융권별 여수신 동향' 표에서 잔액을 읽는다. 게시글 하나 = 잔액기준일 하나.
최근 N개 게시글을 훑어 최근 며칠치 시계열을 만들고, fetch.py 가 아카이브에 누적한다.

수집 항목(잔액, 억원):
  은행:  실세총예금·실세요구불·저축성·금전신탁
  증권:  고객RP (대고객RP매도)

값 단위는 억원(원자료). 대시보드에서 전일 증감을 계산한다.
"""
import re
import ssl
from datetime import date, datetime
from html.parser import HTMLParser
from urllib.parse import urljoin, unquote, urlparse, parse_qs
from urllib.request import Request, urlopen

import openpyxl

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125 Safari/537.36"
RSS_URL = "https://www.bok.or.kr/portal/bbs/P0002018/news.rss?menuNo=200366"

# BOK 표의 항목(공백·괄호 제거 후) → (섹터, 표시명)
TARGET = {
    "실세총예금": "실세총예금",
    "실세요구불": "실세요구불",
    "저축성": "저축성",
    "금전신탁": "금전신탁",
    "대고객RP매도": "고객RP",
    "CMA": "CMA",
}
DATE_SHEET_RE = re.compile(r"^\s*\d{1,2}\.\d{1,2}\s*$")


class BokError(RuntimeError):
    pass


def _get(url):
    ctx = ssl._create_unverified_context()
    req = Request(url, headers={"User-Agent": UA})
    with urlopen(req, timeout=40, context=ctx) as r:
        return r.read()


def _canon(url):
    q = parse_qs(urlparse(url).query).get("nttId", [None])[0]
    return f"https://www.bok.or.kr/portal/bbs/P0002018/view.do?nttId={q}&menuNo=200366" if q else None


def _rss_posts(limit):
    """최근 게시글 [(url, title)] — 최신순."""
    rss = _get(RSS_URL).decode("utf-8", "replace")
    link_re = re.compile(r"<link>\s*(?:<!\[CDATA\[)?\s*(https://www\.bok\.or\.kr/portal/bbs/P0002018/view\.do\?[^\]\s<]+)", re.I)
    title_re = re.compile(r"<title>\s*(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?\s*</title>", re.I | re.S)
    out, seen = [], set()
    for block in re.findall(r"<item>(.*?)</item>", rss, re.I | re.S):
        m = link_re.search(block)
        if not m:
            continue
        c = _canon(m.group(1))
        if not c or c in seen:
            continue
        t = title_re.search(block)
        seen.add(c)
        out.append((c, t.group(1).strip() if t else ""))
        if len(out) >= limit:
            break
    return out


class _Links(HTMLParser):
    def __init__(self):
        super().__init__(); self.hrefs = []
    def handle_starttag(self, tag, attrs):
        if tag.lower() == "a":
            h = dict(attrs).get("href")
            if h:
                self.hrefs.append(h)


def _xlsx_url(page_url):
    """게시글 HTML에서 xlsx 첨부 URL(잔액 우선)."""
    html = _get(page_url).decode("utf-8", "replace")
    p = _Links(); p.feed(html)
    cands = [urljoin(page_url, h) for h in p.hrefs if re.search(r"\.xlsx?\b", unquote(h), re.I)]
    if not cands:
        return None
    bal = [u for u in cands if "잔액" in unquote(u)]
    return (bal or cands)[0]


def _norm(v):
    t = "" if v is None else str(v)
    t = re.sub(r"\d+\)", "", t)
    return re.sub(r"[\s()]+", "", t).replace("*", "")


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    t = str(v).strip().replace(",", "")
    if t in {"", "-", ".."}:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _xldate(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def _parse_sheet(ws):
    """한 시트 → {표시명: (잔액기준일, 잔액값)}. 파싱 불가면 None."""
    sec = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "2. 금융권별 여수신 동향":
                sec = cell.row; break
        if sec:
            break
    if sec is None:
        return None
    hr = sec + 3
    bal_col = bal_date = None
    for col in range(1, ws.max_column + 1):
        d = _xldate(ws.cell(hr, col).value)
        if d and _norm(ws.cell(hr - 1, col).value) == "잔액":
            bal_col, bal_date = col, d
            break
    if bal_col is None:
        return None
    found = {}
    for r in range(sec + 1, ws.max_row + 1):
        key = _norm(ws.cell(r, 3).value)
        if key in TARGET:
            found[TARGET[key]] = (bal_date, _num(ws.cell(r, bal_col).value))
    return found or None


def _parse_workbook(data):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheets = [n for n in wb.sheetnames if n == "일일동향" or DATE_SHEET_RE.match(n)] or [wb.sheetnames[0]]
    # 표시명 → {잔액기준일: 값}
    out = {}
    for name in sheets:
        res = _parse_sheet(wb[name])
        if not res:
            continue
        for disp, (d, v) in res.items():
            if d and v is not None:
                out.setdefault(disp, {})[d] = v
    return out


def fetch(indicator: dict) -> list[dict]:
    posts = int(indicator.get("bok_posts", 12))          # 최근 게시글 수
    collected: dict[str, dict] = {}                       # 표시명 → {date: 억원}
    got = 0
    for page_url, _title in _rss_posts(posts * 2):
        if got >= posts:
            break
        try:
            xu = _xlsx_url(page_url)
            if not xu:
                continue
            parsed = _parse_workbook(_get(xu))
        except Exception as e:
            print(f"  [bok] {page_url} 처리 실패(무시): {e}")
            continue
        if not parsed:
            continue
        for disp, by_date in parsed.items():
            collected.setdefault(disp, {}).update(by_date)
        got += 1

    if not collected:
        raise BokError("BOK 게시글에서 자금흐름 항목을 얻지 못했습니다 (게시판/엑셀 구조 확인)")

    order = ["실세총예금", "실세요구불", "저축성", "금전신탁", "고객RP", "CMA"]
    series = []
    for name in order:
        by_date = collected.get(name)
        if by_date:
            series.append({"name": name, "data": [{"d": d, "v": v} for d, v in sorted(by_date.items())]})
    return series
