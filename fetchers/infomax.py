"""인포맥스 금리 엑셀 수집기 (로컬 파일).

data/인포맥스 금리.xlsx 의 stack 시트를 읽어 시리즈로 변환한다.
API가 아니라 로컬 파일이므로, 엑셀을 새로 내려받아 덮어쓰면 다음 fetch에 반영된다.

시트 구조 (2행=그룹명, 3행=세부명, 4행부터 데이터, A열=일자):
    금리stack       : 국고채권/통안증권/회사채… × 만기(3월이하·3년이하·10년이하…)
    기준금리stack   : 한국:기준금리 / 미국:기준금리 상단 / 한국:BEI 10년 …

indicators.yaml 예:
    params:
      sheet: 금리stack
      groups: ["국고채권", "통안증권"]     # 생략 시 전체
      flat: false                          # true면 그룹명만으로 시리즈명 구성
"""
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
# 파일명이 '인포맥스 금리*.xlsx' 이면 모두 후보로 보고 **가장 최근에 수정된 것**을 쓴다.
# (엑셀이 열려 있어 덮어쓰기가 안 될 때 새 이름으로 떨궈도 자동 인식되도록)
SEARCH_DIRS = [
    ROOT / "data",
    ROOT,
    Path.home() / "Desktop" / "Macro",
]
PATTERN = "인포맥스 금리*.xlsx"


class InfomaxError(RuntimeError):
    pass


def _find_file(indicator: dict) -> Path:
    custom = indicator.get("params", {}).get("path")
    if custom and Path(custom).exists():
        return Path(custom)
    found = []
    for d in SEARCH_DIRS:
        if d.is_dir():
            found += [p for p in d.glob(PATTERN) if p.is_file() and not p.name.startswith("~$")]
    if not found:
        raise InfomaxError(
            "인포맥스 금리 엑셀을 찾을 수 없습니다. "
            f"다음 폴더 중 하나에 '{PATTERN}' 형태로 두세요: "
            + ", ".join(str(d) for d in SEARCH_DIRS))
    return max(found, key=lambda p: p.stat().st_mtime)


def _to_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    m = re.match(r"(\d{4})[-./]?(\d{1,2})[-./]?(\d{1,2})", s)
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else None


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


# 만기 표기 정리: '3년이하(당일)' → '3년', '대표수익률' → ''
def _tenor(s: str) -> str:
    t = re.sub(r"\((당일|적용일)\)", "", str(s)).strip()
    t = re.sub(r"이하$", "", t).strip()
    return "" if t in ("대표수익률", "현재가") else t


def _fetch_flat(ws, p) -> list[dict]:
    """헤더가 한 줄인 시트 (IRS_stack·환율_raw·스프레드 등).

    params:
        header_row: 시리즈명이 있는 행 번호(1부터). 그 다음 행부터 데이터.
        only:  가져올 시리즈명 목록 (생략 시 전체)
        rename: {원본명: 표시명}
    """
    hr = int(p["header_row"])
    only = set(p.get("only") or [])
    rename = p.get("rename") or {}
    rows = list(ws.iter_rows(values_only=True))
    if hr > len(rows):
        raise InfomaxError(f"header_row {hr} 가 시트 범위를 벗어납니다 ({len(rows)}행)")
    hdr = rows[hr - 1]
    cols = []
    for j, v in enumerate(hdr):
        if j == 0 or v is None:
            continue
        nm = str(v).strip()
        if nm in ("일자", "Date"):
            continue
        if only and nm not in only:
            continue
        cols.append((j, rename.get(nm, nm)))
    if not cols:
        raise InfomaxError(f"헤더행 {hr} 에서 대상 열을 찾지 못했습니다 (only={sorted(only)})")
    data = {nm: {} for _j, nm in cols}
    for row in rows[hr:]:
        if not row or row[0] is None:
            continue
        d = _to_date(row[0])
        if not d:
            continue
        for j, nm in cols:
            if j < len(row):
                v = _num(row[j])
                if v is not None:
                    data[nm][d] = v
    return [{"name": n, "data": [{"d": d, "v": v} for d, v in sorted(vals.items())]}
            for n, vals in data.items() if vals]


def fetch(indicator: dict) -> list[dict]:
    path = _find_file(indicator)
    p = indicator.get("params", {})
    sheet = p.get("sheet", "금리stack")
    want = set(p.get("groups") or [])
    flat = bool(p.get("flat"))

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise InfomaxError(f"시트 없음: {sheet} (있는 시트: {wb.sheetnames})")
    ws = wb[sheet]

    if p.get("header_row"):                     # 단일 헤더 레이아웃
        out = _fetch_flat(ws, p)
        if not out:
            raise InfomaxError(f"'{sheet}'에서 데이터를 읽지 못했습니다")
        print(f"  [infomax] {path.name} · {sheet} → 시리즈 {len(out)}개")
        return out

    rows = ws.iter_rows(values_only=True)
    header = []
    for _ in range(3):                     # 1행 공백, 2행 그룹, 3행 세부
        header.append(next(rows, ()))
    grp_row, sub_row = header[1], header[2]

    cur = ""
    cols = []                              # (열idx, 시리즈명)
    for i, (g, s) in enumerate(zip(grp_row, sub_row)):
        if g:
            cur = str(g).strip()
        if i == 0 or not s:
            continue
        if want and cur not in want:
            continue
        name = cur if flat else (f"{cur} {_tenor(s)}".strip() if _tenor(s) else cur)
        cols.append((i, name))
    if not cols:
        raise InfomaxError(f"'{sheet}'에서 대상 열을 찾지 못했습니다 (groups={sorted(want)})")

    data = {name: {} for _, name in cols}
    for row in rows:
        if not row or row[0] is None:
            continue
        d = _to_date(row[0])
        if not d:
            continue
        for i, name in cols:
            if i < len(row):
                v = _num(row[i])
                if v is not None:
                    data[name][d] = v

    out = [{"name": n, "data": [{"d": d, "v": v} for d, v in sorted(vals.items())]}
           for n, vals in data.items() if vals]
    if not out:
        raise InfomaxError(f"'{sheet}'에서 데이터를 읽지 못했습니다")
    print(f"  [infomax] {path.name} · {sheet} → 시리즈 {len(out)}개")
    return out
