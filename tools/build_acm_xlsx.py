"""ACM 기간프리미엄 엑셀 워크북 생성기.

docs/data/kr_yield.json 의 국고채 파금리를 넣고, ACM(Adrian-Crump-Moench 2013) 5단계를
**전부 엑셀 수식으로** 계산하는 워크북을 만든다. 셀을 눌러보면 계산 경로가 다 보인다.

엑셀 구현을 위해 파이썬판(fetchers/acm.py)과 세 군데를 바꿨다. 모두 영향이 작다는 걸
검증 시트에서 수치로 보여준다.

  1) 상태변수: 주성분 5개 → 관측가능 3팩터(레벨/기울기/곡률). 엑셀에 고유값 분해가 없다.
  2) 무이표 커브: 부트스트랩 생략, 파금리 그대로. 10년 기준 0.7bp 차이.
  3) BRW 편의보정 Phi는 부트스트랩이 필요해 수식으로 못 만든다. 파이썬으로 계산한 3x3을
     상수로 넣고 스위치로 전환한다. 그 아래 계산은 전부 선택된 Phi를 타고 다시 흐른다.

사용법:
    python tools/build_acm_xlsx.py            # Macro 폴더로 저장
    python tools/build_acm_xlsx.py 경로.xlsx
"""
import json
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "docs" / "data"

TENORS = [("3월", 3), ("6월", 6), ("9월", 9), ("1년", 12), ("1.5년", 18), ("2년", 24),
          ("2.5년", 30), ("3년", 36), ("4년", 48), ("5년", 60), ("7년", 84), ("10년", 120)]
SRC_M = [m for _n, m in TENORS]
RX_N = [6, 12, 24, 36, 48, 60, 84, 120]        # 초과수익률 대상 만기(개월) — 파이썬판과 동일
# 필요한 만기: 팩터용(1·3·24·120) + 결과용(36·120) + 초과수익률용(n 과 n−1)
MATS = sorted({1, 3, 24, 36, 120} | set(RX_N) | {n - 1 for n in RX_N})
NMAX = 120
K = 3
R0 = 5

F = "Arial"
BLUE = Font(name=F, size=10, color="0000FF")
BLK = Font(name=F, size=10)
GRN = Font(name=F, size=10, color="008000")
HDR = Font(name=F, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=F, size=13, bold=True)
SUB = Font(name=F, size=10, bold=True)
NOTE = Font(name=F, size=9, color="808080")
RED = Font(name=F, size=10, bold=True, color="C0392B")
HFILL = PatternFill("solid", fgColor="1A4FA0")
YFILL = PatternFill("solid", fgColor="FFFF00")
BOX = Border(*[Side("thin", color="D7DDE8")] * 4)
CEN = Alignment(horizontal="center", vertical="center")


# ───────────────────────────── 셀 블록 헬퍼
class Block:
    """워크시트 위의 행렬 블록. 참조 문자열을 안전하게 만들어 준다."""

    def __init__(self, sheet, row, col, rows, cols):
        self.sheet, self.row, self.col, self.rows, self.cols = sheet, row, col, rows, cols

    def c(self, i, j=0, q=False):
        ref = f"${CL(self.col + j)}${self.row + i}"
        return f"{self.sheet}!{ref}" if q else ref

    def rng(self, q=False):
        ref = (f"${CL(self.col)}${self.row}"
               f":${CL(self.col + self.cols - 1)}${self.row + self.rows - 1}")
        return f"{self.sheet}!{ref}" if q else ref


def put(ws, cell, val, font=BLK, fmt=None):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt:
        c.number_format = fmt
    return c


def head(ws, row, labels, start=1):
    for i, t in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=t)
        c.font, c.fill, c.alignment, c.border = HDR, HFILL, CEN, BOX


def block(ws, row, col, rows, cols, fn, fmt="0.00000000", font=BLK, rlab=None, clab=None):
    """(row, col) 을 데이터 좌상단으로 rows x cols 블록을 쓴다. 라벨은 그 위/왼쪽."""
    if clab:
        for j, t in enumerate(clab):
            put(ws, f"{CL(col + j)}{row - 1}", t, SUB).alignment = CEN
    if rlab:
        for i, t in enumerate(rlab):
            put(ws, f"{CL(col - 1)}{row + i}", t, SUB)
    for i in range(rows):
        for j in range(cols):
            put(ws, f"{CL(col + j)}{row + i}", fn(i, j), font, fmt).border = BOX
    return Block(ws.title, row, col, rows, cols)


# ───────────────────────────── 데이터
def load_panel():
    payload = json.loads((DATA / "kr_yield.json").read_text(encoding="utf-8"))
    names = [f"국고채권 {n}" for n, _ in TENORS]
    got = {s["name"]: {p["d"]: p["v"] for p in s["data"]}
           for s in payload["series"] if s["name"] in names}
    miss = [n for n in names if n not in got]
    if miss:
        sys.exit(f"kr_yield 에 없는 만기: {miss}")
    common = sorted(set.intersection(*[set(v) for v in got.values()]))
    last = {}
    for d in common:
        last[d[:7]] = d
    months = sorted(last)
    par = np.array([[got[n][last[m]] for n in names] for m in months])
    return months, par


def _var_ols(V):
    reg = np.column_stack([np.ones(len(V) - 1), V[:-1]])
    cf = np.linalg.lstsq(reg, V[1:], rcond=None)[0]
    return cf[0], cf[1:].T, V[1:] - reg @ cf


def reference(par):
    """엑셀과 동일한 모형(관측 3팩터·파금리)을 파이썬으로 풀어 검증 기준값을 만든다."""
    mg = np.arange(1, NMAX + 1)                  # 만기 격자 — SRC_M 과 같은 '개월' 단위
    Z = np.array([np.interp(mg, SRC_M, par[i]) / 100.0 for i in range(len(par))])
    lp = -(np.arange(1, NMAX + 1) * Z / 12.0)
    r1 = -lp[:, 0]
    Xr = np.column_stack([Z[:, 119], Z[:, 119] - Z[:, 2], 2 * Z[:, 23] - Z[:, 2] - Z[:, 119]])
    X = Xr - Xr.mean(0)
    x0, x1 = X[:-1], X[1:]
    _mu, phi, res = _var_ols(X)

    def brw(phi_ols, resid, cap=0.999, nb=2000, ni=40, seed=7):
        rng = np.random.default_rng(seed)
        T = len(X) - 1
        pt = phi_ols.copy()
        for _ in range(ni):
            acc = np.zeros_like(pt)
            for _ in range(nb):
                vb = resid[rng.integers(0, T, T)]
                sim = np.empty((T + 1, K))
                sim[0] = X[0]
                for t in range(T):
                    sim[t + 1] = pt @ sim[t] + vb[t]
                acc += _var_ols(sim)[1]
            step = phi_ols - acc / nb
            cand = pt + step
            if max(abs(np.linalg.eigvals(cand))) >= cap:
                lo, hi = 0.0, 1.0
                for _ in range(40):
                    mid = (lo + hi) / 2
                    if max(abs(np.linalg.eigvals(pt + mid * step))) < cap:
                        lo = mid
                    else:
                        hi = mid
                return pt + lo * step
            if np.max(np.abs(step)) < 1e-7:
                return cand
            pt = cand
        return pt

    phi_b = brw(phi, res)
    out = {}
    for tag, ph in (("OLS", phi), ("BRW", phi_b)):
        rr = x1 - x0 @ ph.T
        mm = x1.mean(0) - ph @ x0.mean(0)
        rr = rr - rr.mean(0)
        T = len(x0)
        rx = np.column_stack([lp[1:, n - 2] - lp[:-1, n - 1] - r1[:-1] for n in RX_N])
        W = np.column_stack([np.ones(T), rr, x0])
        G = np.linalg.lstsq(W, rx, rcond=None)[0]
        a, beta, c = G[0], G[1:1 + K], G[1 + K:]
        sig2 = float(((rx - W @ G) ** 2).sum() / (T * len(RX_N)))
        sig = rr.T @ rr / T
        bb = beta @ beta.T
        rhs = a + 0.5 * (np.array([beta[:, i] @ sig @ beta[:, i]
                                   for i in range(len(RX_N))]) + sig2)
        l0 = np.linalg.solve(bb, beta @ rhs)
        l1 = np.linalg.solve(bb, beta @ c.T)
        dz = np.linalg.lstsq(np.column_stack([np.ones(len(X)), X]), r1, rcond=None)[0]
        d0, d1 = dz[0], dz[1:]

        def rec(L0, L1):
            A = np.zeros(NMAX + 1)
            B = np.zeros((NMAX + 1, K))
            for n in range(1, NMAX + 1):
                bp = B[n - 1]
                A[n] = A[n - 1] + bp @ (mm - L0) + 0.5 * (bp @ sig @ bp + sig2) - d0
                B[n] = bp @ (ph - L1) - d1
            return A, B

        Af, Bf = rec(l0, l1)
        Aq, Bq = rec(np.zeros(K), np.zeros((K, K)))
        d = {}
        for n, lab in ((36, "3년"), (120, "10년")):
            fit = -(Af[n] + X @ Bf[n]) / n * 1200
            rn = -(Aq[n] + X @ Bq[n]) / n * 1200
            d[lab] = dict(rmse_bp=float(np.sqrt(((fit - Z[:, n - 1] * 100) ** 2).mean()) * 100),
                          rn=float(rn[-1]), tp=float((fit - rn)[-1]),
                          tp_mean=float((fit - rn).mean()), fit=float(fit[-1]))
        out[tag] = d
    return phi, phi_b, out


def interp_formula(row, m):
    """원자료의 파금리를 목표 만기 m(개월)으로 선형보간.

    관측 구간(3~120개월) 밖은 외삽하지 않고 끝값으로 고정한다. 1개월 금리는 3개월 금리와
    같게 두는 셈인데, numpy.interp 의 동작(클램프)과 일치시켜 파이썬판과 결과가 어긋나지
    않게 하기 위해서다. 3개월 밑을 외삽하면 근거 없는 기울기가 단기금리식에 섞여 들어간다.
    """
    if m in SRC_M:
        return f"=원자료!{CL(2 + SRC_M.index(m))}{row}"
    if m < SRC_M[0]:
        return f"=원자료!{CL(2)}{row}"
    if m > SRC_M[-1]:
        return f"=원자료!{CL(1 + len(SRC_M))}{row}"
    i = next(j for j in range(len(SRC_M) - 1) if SRC_M[j] <= m <= SRC_M[j + 1])
    lo, hi = SRC_M[i], SRC_M[i + 1]
    a, b = CL(2 + i), CL(3 + i)
    return f"=원자료!{a}{row}+(원자료!{b}{row}-원자료!{a}{row})*({m}-{lo})/({hi}-{lo})"


# ───────────────────────────── 워크북
def build(out_path):
    months, par = load_panel()
    T1 = len(months)
    T = T1 - 1
    phi_ols, phi_brw, ref = reference(par)
    end = R0 + T1 - 1
    rend = R0 + T - 1
    e_ols = float(max(abs(np.linalg.eigvals(phi_ols))))
    e_brw = float(max(abs(np.linalg.eigvals(phi_brw))))
    wb = Workbook()

    # ── 읽기
    ws = wb.active
    ws.title = "읽기"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 108
    put(ws, "B2", "ACM 기간프리미엄 — 엑셀 구현", TITLE)
    lines = [
        ("h", "무엇인가"),
        ("p", "Adrian-Crump-Moench(2013). 아핀 기간구조모형을 최대우도 대신 선형회귀 5단계로 "
              "추정한다. NY Fed가 이 방식으로 미국 기간프리미엄을 매일 공표한다."),
        ("", ""),
        ("p", "① 상태변수 X (수익률 곡선 팩터)"),
        ("p", "② VAR(1):  X(t+1) = mu + Phi·X(t) + v(t+1)"),
        ("p", "③ 초과수익률  rx = p(t+1, n−1) − p(t, n) − r(t)"),
        ("p", "④ rx ~ [1, v(t+1), X(t)] 회귀 → 위험가격 lambda0, lambda1"),
        ("p", "⑤ 아핀 점화식으로 곡선 복원. lambda = 0 으로 한 번 더 → 위험중립금리"),
        ("p", "     기간프리미엄 = 실제금리 − 위험중립금리"),
        ("", ""),
        ("p", "위험중립금리는 투자자가 위험보상을 요구하지 않는다면 붙었을 금리 — 순수 정책금리 "
              "기대 경로다. 기간프리미엄은 그 위에 얹힌 위험 보상이다."),
        ("", ""),
        ("h", "계산 흐름 (시트 순서)"),
        ("p", "원자료 → 무이표금리 → 상태변수 → VAR → 초과수익률 → 위험가격 → 점화식 → 결과 → 검증"),
        ("", ""),
        ("h", "색 규칙"),
        ("p", "파란 글씨 = 직접 넣은 값(원자료, BRW 행렬).   초록 글씨 = 다른 시트 참조.   "
              "검은 글씨 = 이 시트의 수식.   노란 칸 = 사용자가 바꾸는 스위치."),
        ("", ""),
        ("h", "스위치 하나만 기억하세요"),
        ("p", "VAR 시트 D2 셀에서 OLS / BRW 를 고릅니다. 바꾸면 위험가격·점화식·결과가 전부 다시 "
              "계산됩니다. 기본값은 BRW(편의보정)입니다."),
        ("", ""),
        ("h", "왜 편의보정이 필요한가"),
        ("p", f"표본이 {T1}개월뿐이라 VAR 지속성 Phi가 하향 편의된다. 그러면 위험중립금리가 너무 "
              "빨리 평균회귀해 기간프리미엄이 과대추정된다 (Bauer-Rudebusch-Wu 2012). 실제로 Phi "
              f"최대고유값이 OLS {e_ols:.4f} → 보정 {e_brw:.4f} 로 올라가고, 10년 기간프리미엄은 "
              f"{ref['OLS']['10년']['tp']:+.2f}%p → {ref['BRW']['10년']['tp']:+.2f}%p 로 절반이 된다."),
        ("", ""),
        ("h", "파이썬판(fetchers/acm.py)과 다른 점 — 셋 다 영향이 작다 (검증 시트 참조)"),
        ("p", "① 상태변수를 주성분 5개가 아니라 관측가능 3팩터(레벨·기울기·곡률)로 썼다. 엑셀에 "
              "고유값 분해가 없기 때문이다. 기간프리미엄 차이는 3년 2bp / 10년 12bp."),
        ("p", "② 무이표 커브 부트스트랩을 생략하고 파금리를 그대로 썼다. 10년 기준 0.7bp 차이."),
        ("p", "③ BRW 편의보정 Phi는 부트스트랩(2000회 재표본 × 40회 반복)이 필요해 수식으로 만들 수 "
              "없다. 파이썬으로 계산한 3x3 행렬을 VAR 시트에 상수로 넣었다."),
        ("", ""),
        ("h", "읽는 법"),
        ("p", "기간프리미엄의 절대 수준보다 방향·변화분을 신뢰할 것. 아핀 모형의 프리미엄 수준은 "
              "표본 길이와 팩터 수에 민감하다."),
        ("", ""),
        ("p", f"데이터: docs/data/kr_yield.json (인포맥스) · {months[0]} ~ {months[-1]} · {T1}개월"),
    ]
    r = 4
    for kind, txt in lines:
        c = put(ws, f"B{r}", txt, SUB if kind == "h" else BLK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 13 if not txt else (14 if len(txt) < 95 else 15 + 13 * (len(txt) // 95))
        r += 1

    # ── 원자료
    ws = wb.create_sheet("원자료")
    put(ws, "A1", "국고채 파금리 (월말 영업일, %)", TITLE)
    put(ws, "A2", "출처: docs/data/kr_yield.json (인포맥스 금리STACK). 파란 글씨가 입력값이며, "
                  "여기만 갱신하면 워크북 전체가 다시 계산됩니다.", NOTE)
    head(ws, R0 - 2, ["월"] + [n for n, _ in TENORS])
    head(ws, R0 - 1, ["만기(개월)"] + [str(m) for _n, m in TENORS])
    ws.column_dimensions["A"].width = 10
    for i in range(len(TENORS)):
        ws.column_dimensions[CL(2 + i)].width = 9
    for i, m in enumerate(months):
        put(ws, f"A{R0 + i}", m).alignment = CEN
        for j in range(len(TENORS)):
            put(ws, f"{CL(2 + j)}{R0 + i}", round(float(par[i][j]), 4), BLUE, "0.000")
    ws.freeze_panes = f"B{R0}"

    # ── 무이표금리
    ws = wb.create_sheet("무이표금리")
    put(ws, "A1", "만기별 금리와 로그가격", TITLE)
    put(ws, "A2", "원자료 파금리를 필요한 만기(개월)로 선형보간. 관측 구간(3~120개월) 밖은 외삽하지 않고 "
                  "끝값 고정 — 1개월 금리 = 3개월 금리. 파금리를 무이표금리 대용으로 쓴다 (10년 기준 0.7bp 차이).",
        NOTE)
    pcol = 2 + len(MATS) + 1
    rcol = pcol + len(MATS) + 1
    put(ws, f"B{R0 - 3}", "금리 y(n) — 연율 %", SUB)
    put(ws, f"{CL(pcol)}{R0 - 3}", "로그가격 p(n) = −n × y(n) / 1200", SUB)
    put(ws, f"{CL(rcol)}{R0 - 3}", "1개월 무위험수익률 r1 = −p(1)", SUB)
    head(ws, R0 - 2, ["월"] + [f"y{m}" for m in MATS])
    head(ws, R0 - 1, ["만기(개월)"] + [str(m) for m in MATS])
    head(ws, R0 - 2, [f"p{m}" for m in MATS], start=pcol)
    head(ws, R0 - 1, [str(m) for m in MATS], start=pcol)
    head(ws, R0 - 2, ["r1"], start=rcol)
    head(ws, R0 - 1, ["1"], start=rcol)
    ws.column_dimensions["A"].width = 10
    for i in range(T1):
        rw = R0 + i
        put(ws, f"A{rw}", f"=원자료!A{rw}", GRN).alignment = CEN
        for j, m in enumerate(MATS):
            put(ws, f"{CL(2 + j)}{rw}", interp_formula(rw, m), BLK, "0.0000")
            put(ws, f"{CL(pcol + j)}{rw}", f"=-{m}*{CL(2 + j)}{rw}/1200", BLK, "0.00000000")
        put(ws, f"{CL(rcol)}{rw}", f"=-{CL(pcol)}{rw}", BLK, "0.00000000")
    ws.freeze_panes = f"B{R0}"
    YC = {m: CL(2 + j) for j, m in enumerate(MATS)}
    PC = {m: CL(pcol + j) for j, m in enumerate(MATS)}
    RC = CL(rcol)

    # ── 상태변수
    ws = wb.create_sheet("상태변수")
    put(ws, "A1", "① 상태변수 X — 관측가능 3팩터 (소수, %가 아님)", TITLE)
    put(ws, "A2", "레벨 = 10년 · 기울기 = 10년 − 3개월 · 곡률 = 2×2년 − 3개월 − 10년. "
                  "100으로 나눠 소수로 쓰고, 표본평균을 빼 중심화한다 (ACM은 평균 0 상태변수를 쓴다).", NOTE)
    put(ws, "A3", "표본평균 →", SUB)
    head(ws, R0 - 1, ["월", "레벨(원)", "기울기(원)", "곡률(원)", "X1 레벨", "X2 기울기", "X3 곡률"])
    ws.column_dimensions["A"].width = 10
    for j in range(6):
        ws.column_dimensions[CL(2 + j)].width = 13
    for j in range(3):
        put(ws, f"{CL(2 + j)}3", f"=AVERAGE({CL(2 + j)}{R0}:{CL(2 + j)}{end})", BLK, "0.00000000")
    for i in range(T1):
        rw = R0 + i
        put(ws, f"A{rw}", f"=무이표금리!A{rw}", GRN).alignment = CEN
        put(ws, f"B{rw}", f"=무이표금리!{YC[120]}{rw}/100", GRN, "0.00000000")
        put(ws, f"C{rw}", f"=(무이표금리!{YC[120]}{rw}-무이표금리!{YC[3]}{rw})/100", GRN, "0.00000000")
        put(ws, f"D{rw}",
            f"=(2*무이표금리!{YC[24]}{rw}-무이표금리!{YC[3]}{rw}-무이표금리!{YC[120]}{rw})/100",
            GRN, "0.00000000")
        for j in range(3):
            put(ws, f"{CL(5 + j)}{rw}", f"={CL(2 + j)}{rw}-{CL(2 + j)}$3", BLK, "0.00000000")
    ws.freeze_panes = f"B{R0}"
    XC = ["E", "F", "G"]

    # ── VAR
    ws = wb.create_sheet("VAR")
    put(ws, "A1", "② VAR(1):  X(t+1) = mu + Phi·X(t) + v(t+1)", TITLE)
    put(ws, "B2", "Phi 선택 →", SUB)
    sw = put(ws, "D2", "BRW", Font(name=F, size=11, bold=True))
    sw.fill, sw.alignment, sw.border = YFILL, CEN, BOX
    dv = DataValidation(type="list", formula1='"OLS,BRW"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(sw)
    put(ws, "E2", "← OLS = 이 시트에서 회귀한 값 / BRW = 부트스트랩 편의보정(파이썬 계산). "
                  "바꾸면 아래 모든 계산과 결과 시트가 다시 흐릅니다.", NOTE)
    head(ws, R0 - 1, ["월(t)", "상수", "X1(t)", "X2(t)", "X3(t)",
                      "X1(t+1)", "X2(t+1)", "X3(t+1)", "v1(t+1)", "v2(t+1)", "v3(t+1)"], start=2)
    ws.column_dimensions["A"].width = 3
    for j in range(11):
        ws.column_dimensions[CL(2 + j)].width = 11
    REG = ["C", "D", "E", "F"]          # 상수, X1(t), X2(t), X3(t)
    DEP = ["G", "H", "I"]               # X1(t+1)..X3(t+1)
    VV = ["J", "K", "L"]                # 잔차
    for i in range(T):
        rw = R0 + i
        put(ws, f"B{rw}", f"=상태변수!A{rw}", GRN).alignment = CEN
        put(ws, f"C{rw}", 1, BLK, "0")
        for j in range(3):
            put(ws, f"{CL(4 + j)}{rw}", f"=상태변수!{XC[j]}{rw}", GRN, "0.00000000")
            put(ws, f"{CL(7 + j)}{rw}", f"=상태변수!{XC[j]}{rw + 1}", GRN, "0.00000000")

    put(ws, "N3", "정규방정식  X'X (4x4)", SUB)
    XtX = block(ws, 5, 16, 4, 4,
                lambda i, j: f"=SUMPRODUCT({REG[i]}${R0}:{REG[i]}${rend},"
                             f"{REG[j]}${R0}:{REG[j]}${rend})",
                rlab=["const", "X1", "X2", "X3"], clab=["const", "X1", "X2", "X3"])
    put(ws, "U3", "X'y (4x3)", SUB)
    Xty = block(ws, 5, 22, 4, 3,
                lambda i, j: f"=SUMPRODUCT({REG[i]}${R0}:{REG[i]}${rend},"
                             f"{DEP[j]}${R0}:{DEP[j]}${rend})",
                clab=["→X1", "→X2", "→X3"])
    put(ws, "N11", "회귀계수 (4x3) = MINVERSE(X'X) × X'y", SUB)
    CFb = block(ws, 13, 16, 4, 3,
                lambda i, j: f"=INDEX(MMULT(MINVERSE({XtX.rng()}),{Xty.rng()}),{i + 1},{j + 1})",
                rlab=["mu", "X1", "X2", "X3"], clab=["→X1", "→X2", "→X3"])
    put(ws, "N18", "Phi_OLS   (행 i = X_i(t+1), 열 j = X_j(t))", SUB)
    POLS = block(ws, 20, 16, 3, 3, lambda i, j: f"={CFb.c(1 + j, i)}",
                 rlab=["X1", "X2", "X3"], clab=["X1", "X2", "X3"])
    put(ws, "N24", "Phi_BRW   (파이썬 부트스트랩 편의보정 — 상수)", SUB)
    put(ws, "N25", "Bauer-Rudebusch-Wu(2012) indirect inference · 재표본 2000회 × 반복 40회 · "
                   f"정상성 상한 0.999 → 최대고유값 {e_brw:.4f}", NOTE)
    PBRW = block(ws, 27, 16, 3, 3, lambda i, j: round(float(phi_brw[i][j]), 10), font=BLUE,
                 rlab=["X1", "X2", "X3"], clab=["X1", "X2", "X3"])
    put(ws, "N31", "Phi_사용   = D2 스위치에 따라 선택", SUB)
    PHI = block(ws, 33, 16, 3, 3,
                lambda i, j: f'=IF($D$2="BRW",{PBRW.c(i, j)},{POLS.c(i, j)})',
                rlab=["X1", "X2", "X3"], clab=["X1", "X2", "X3"])
    put(ws, "N37", "mu_사용   = 평균(X(t+1)) − Phi × 평균(X(t))", SUB)
    MU = block(ws, 39, 16, 3, 1,
               lambda i, j: "=AVERAGE({d}${a}:{d}${b})-({t})".format(
                   d=DEP[i], a=R0, b=rend,
                   t="+".join(f"{PHI.c(i, k)}*AVERAGE({REG[1 + k]}${R0}:{REG[1 + k]}${rend})"
                              for k in range(3))),
               rlab=["X1", "X2", "X3"], clab=["mu"])
    for i in range(T):
        rw = R0 + i
        for j in range(3):
            terms = "+".join(f"{PHI.c(j, k)}*{REG[1 + k]}{rw}" for k in range(3))
            put(ws, f"{VV[j]}{rw}", f"={DEP[j]}{rw}-{MU.c(j)}-({terms})", BLK, "0.00000000")
    put(ws, "N43", f"Sigma = v'v / T   (T = {T})", SUB)
    SIG = block(ws, 45, 16, 3, 3,
                lambda i, j: f"=SUMPRODUCT({VV[i]}${R0}:{VV[i]}${rend},"
                             f"{VV[j]}${R0}:{VV[j]}${rend})/{T}",
                rlab=["v1", "v2", "v3"], clab=["v1", "v2", "v3"])
    put(ws, "N49", "Phi 최대고유값 (참고, 파이썬 계산)", SUB)
    put(ws, "N50", f"OLS {e_ols:.4f}  (반감기 {np.log(0.5) / np.log(e_ols):.1f}개월)", BLUE)
    put(ws, "N51", f"BRW {e_brw:.4f}  (반감기 {np.log(0.5) / np.log(e_brw):.0f}개월)", BLUE)
    ws.freeze_panes = f"C{R0}"

    # ── 초과수익률
    ws = wb.create_sheet("초과수익률")
    put(ws, "A1", "③ 초과수익률  rx(t→t+1, n) = p(t+1, n−1) − p(t, n) − r1(t)", TITLE)
    put(ws, "A2", "n개월 채권을 사서 한 달 뒤 (n−1)개월 채권으로 팔 때, 무위험수익률을 넘는 로그수익률.", NOTE)
    head(ws, R0 - 1, ["월(t)"] + [f"rx {n}개월" for n in RX_N])
    ws.column_dimensions["A"].width = 10
    for j in range(len(RX_N)):
        ws.column_dimensions[CL(2 + j)].width = 13
    for i in range(T):
        rw = R0 + i
        put(ws, f"A{rw}", f"=무이표금리!A{rw}", GRN).alignment = CEN
        for j, n in enumerate(RX_N):
            put(ws, f"{CL(2 + j)}{rw}",
                f"=무이표금리!{PC[n - 1]}{rw + 1}-무이표금리!{PC[n]}{rw}-무이표금리!{RC}{rw}",
                BLK, "0.00000000")
    ws.freeze_panes = f"B{R0}"

    # ── 위험가격
    ws = wb.create_sheet("위험가격")
    put(ws, "A1", "④ rx ~ [1, v(t+1), X(t)] 회귀 → 위험가격 lambda0, lambda1", TITLE)
    put(ws, "A2", "v 계수(beta)는 위험 노출도, X(t) 계수(c)는 위험가격의 상태 의존성을 잡는다.", NOTE)
    NR = len(RX_N)
    W = [CL(2 + i) for i in range(7)]            # B..H : 상수, v1..v3, X1..X3(t)
    RXCOL = 10                                   # J.. : rx
    RXc = [CL(RXCOL + j) for j in range(NR)]
    ECOL = RXCOL + NR + 1                        # 회귀 잔차 e
    SCOL = ECOL + NR + 1                         # 단기금리 회귀 보조열
    MCOL = SCOL + 5 + 2                          # 행렬 블록 시작 (라벨은 MCOL-1)
    ML = CL(MCOL - 1)
    head(ws, R0 - 1, ["상수", "v1", "v2", "v3", "X1(t)", "X2(t)", "X3(t)"], start=2)
    head(ws, R0 - 1, [f"rx{n}" for n in RX_N], start=RXCOL)
    head(ws, R0 - 1, [f"e{n}" for n in RX_N], start=ECOL)
    head(ws, R0 - 1, ["상수", "X1", "X2", "X3", "r1"], start=SCOL)
    ws.column_dimensions["A"].width = 3
    for j in range(MCOL + 12):
        ws.column_dimensions[CL(2 + j)].width = 11
    for i in range(T):
        rw = R0 + i
        put(ws, f"B{rw}", 1, BLK, "0")
        for j in range(3):
            put(ws, f"{W[1 + j]}{rw}", f"=VAR!{VV[j]}{rw}", GRN, "0.00000000")
            put(ws, f"{W[4 + j]}{rw}", f"=VAR!{REG[1 + j]}{rw}", GRN, "0.00000000")
        for j in range(NR):
            put(ws, f"{RXc[j]}{rw}", f"=초과수익률!{CL(2 + j)}{rw}", GRN, "0.00000000")
    wl = ["1", "v1", "v2", "v3", "X1", "X2", "X3"]
    put(ws, f"{ML}3", "W'W (7x7)", SUB)
    WtW = block(ws, 5, MCOL, 7, 7,
                lambda i, j: f"=SUMPRODUCT({W[i]}${R0}:{W[i]}${rend},{W[j]}${R0}:{W[j]}${rend})",
                rlab=wl, clab=wl)
    put(ws, f"{CL(MCOL + 9)}3", f"W'rx (7x{NR})", SUB)
    Wtr = block(ws, 5, MCOL + 9, 7, NR,
                lambda i, j: f"=SUMPRODUCT({W[i]}${R0}:{W[i]}${rend},"
                             f"{RXc[j]}${R0}:{RXc[j]}${rend})",
                clab=[f"rx{n}" for n in RX_N])
    put(ws, f"{ML}14", f"회귀계수 G (7x{NR}) = MINVERSE(W'W) × W'rx      "
                       "1행 = a · 2~4행 = beta · 5~7행 = c", SUB)
    G = block(ws, 16, MCOL, 7, NR,
              lambda i, j: f"=INDEX(MMULT(MINVERSE({WtW.rng()}),{Wtr.rng()}),{i + 1},{j + 1})",
              rlab=["a", "beta1", "beta2", "beta3", "c1", "c2", "c3"],
              clab=[f"rx{n}" for n in RX_N])
    A_ = lambda j: G.c(0, j)
    BE = lambda i, j: G.c(1 + i, j)
    CC = lambda i, j: G.c(4 + i, j)
    for i in range(T):
        rw = R0 + i
        for j in range(NR):
            terms = "+".join(f"{W[k]}{rw}*{G.c(k, j)}" for k in range(7))
            put(ws, f"{CL(ECOL + j)}{rw}", f"={RXc[j]}{rw}-({terms})", BLK, "0.00000000")
    put(ws, f"{ML}24", f"sigma^2 = 잔차제곱합 / (T × 만기수)      잔차 e 는 "
                       f"{CL(ECOL)}:{CL(ECOL + NR - 1)} 열", SUB)
    S2 = block(ws, 25, MCOL, 1, 1,
               lambda i, j: "=(" + "+".join(
                   f"SUMPRODUCT({CL(ECOL + k)}${R0}:{CL(ECOL + k)}${rend},"
                   f"{CL(ECOL + k)}${R0}:{CL(ECOL + k)}${rend})" for k in range(NR))
               + f")/({T}*{NR})",
               fmt="0.000000E+00", rlab=["sigma^2"])
    put(ws, f"{ML}28", "만기별  beta'·Sigma·beta", SUB)
    BSB = block(ws, 29, MCOL, 1, NR,
                lambda i, j: "=" + "+".join(
                    f"{BE(p, j)}*{SIG.c(p, q, q=True)}*{BE(q, j)}"
                    for p in range(3) for q in range(3)),
                rlab=["b'Sb"], clab=[f"rx{n}" for n in RX_N])
    put(ws, f"{ML}31", "rhs = a + 0.5 × (beta'·Sigma·beta + sigma^2)", SUB)
    RHS = block(ws, 32, MCOL, 1, NR,
                lambda i, j: f"={A_(j)}+0.5*({BSB.c(0, j)}+{S2.c(0, 0)})", rlab=["rhs"])
    put(ws, f"{ML}35", "BB = beta × beta' (3x3)", SUB)
    BB = block(ws, 37, MCOL, 3, 3,
               lambda i, j: "=" + "+".join(f"{BE(i, k)}*{BE(j, k)}" for k in range(NR)),
               rlab=["1", "2", "3"], clab=["1", "2", "3"])
    put(ws, f"{CL(MCOL + 5)}35", "beta × rhs (3x1)", SUB)
    BR = block(ws, 37, MCOL + 5, 3, 1,
               lambda i, j: "=" + "+".join(f"{BE(i, k)}*{RHS.c(0, k)}" for k in range(NR)),
               clab=["v"])
    put(ws, f"{CL(MCOL + 8)}35", "beta × c' (3x3)", SUB)
    BC = block(ws, 37, MCOL + 8, 3, 3,
               lambda i, j: "=" + "+".join(f"{BE(i, k)}*{CC(j, k)}" for k in range(NR)),
               clab=["1", "2", "3"])
    put(ws, f"{ML}41", "lambda0 = MINVERSE(BB) × (beta×rhs)", SUB)
    L0 = block(ws, 43, MCOL, 3, 1,
               lambda i, j: f"=INDEX(MMULT(MINVERSE({BB.rng()}),{BR.rng()}),{i + 1},1)",
               rlab=["1", "2", "3"], clab=["lambda0"])
    put(ws, f"{CL(MCOL + 5)}41", "lambda1 = MINVERSE(BB) × (beta×c')", SUB)
    L1 = block(ws, 43, MCOL + 5, 3, 3,
               lambda i, j: f"=INDEX(MMULT(MINVERSE({BB.rng()}),{BC.rng()}),{i + 1},{j + 1})",
               clab=["1", "2", "3"])
    SW_ = [CL(SCOL + i) for i in range(4)]
    SY = CL(SCOL + 4)
    for i in range(T1):
        rw = R0 + i
        put(ws, f"{SW_[0]}{rw}", 1, BLK, "0")
        for j in range(3):
            put(ws, f"{SW_[1 + j]}{rw}", f"=상태변수!{XC[j]}{rw}", GRN, "0.00000000")
        put(ws, f"{SY}{rw}", f"=무이표금리!{RC}{rw}", GRN, "0.00000000")
    put(ws, f"{ML}48", f"단기금리식  r1(t) = d0 + d1'·X(t)   (전체 {T1}개월, 보조열 "
                       f"{CL(SCOL)}:{CL(SCOL + 4)})", SUB)
    DtD = block(ws, 50, MCOL, 4, 4,
                lambda i, j: f"=SUMPRODUCT({SW_[i]}${R0}:{SW_[i]}${end},"
                             f"{SW_[j]}${R0}:{SW_[j]}${end})",
                rlab=["const", "X1", "X2", "X3"], clab=["const", "X1", "X2", "X3"])
    Dty = block(ws, 50, MCOL + 6, 4, 1,
                lambda i, j: f"=SUMPRODUCT({SW_[i]}${R0}:{SW_[i]}${end},{SY}${R0}:{SY}${end})",
                clab=["X'y"])
    DC = block(ws, 56, MCOL, 4, 1,
               lambda i, j: f"=INDEX(MMULT(MINVERSE({DtD.rng()}),{Dty.rng()}),{i + 1},1)",
               rlab=["d0", "d1_1", "d1_2", "d1_3"], clab=["계수"])
    ws.freeze_panes = f"B{R0}"

    # ── 점화식
    ws = wb.create_sheet("점화식")
    put(ws, "A1", "⑤ 아핀 점화식", TITLE)
    put(ws, "A2", "A(n) = A(n−1) + B(n−1)'·(mu − lambda0) + 0.5×(B(n−1)'·Sigma·B(n−1) + sigma^2) − d0", NOTE)
    put(ws, "A3", "B(n)' = B(n−1)'·(Phi − lambda1) − d1'          위험중립(~)은 lambda0 = lambda1 = 0", NOTE)
    head(ws, R0 - 1, ["n(개월)", "A(n)", "B1(n)", "B2(n)", "B3(n)", "",
                      "A~(n)", "B1~(n)", "B2~(n)", "B3~(n)"])
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["F"].width = 3
    for j in [1, 2, 3, 4, 6, 7, 8, 9]:
        ws.column_dimensions[CL(2 + j)].width = 13
    put(ws, f"A{R0}", 0, BLK, "0")
    for c in ("B", "C", "D", "E", "G", "H", "I", "J"):
        put(ws, f"{c}{R0}", 0, BLK, "0.00000000")
    for n in range(1, NMAX + 1):
        rw, pv = R0 + n, R0 + n - 1
        put(ws, f"A{rw}", n, BLK, "0")
        for Acol, Bc, use_l in (("B", ["C", "D", "E"], True), ("G", ["H", "I", "J"], False)):
            mu_t = "+".join(f"{Bc[i]}{pv}*({MU.c(i, 0, q=True)}"
                            + (f"-{L0.c(i, 0, q=True)})" if use_l else ")") for i in range(3))
            quad = "+".join(f"{Bc[i]}{pv}*{SIG.c(i, j, q=True)}*{Bc[j]}{pv}"
                            for i in range(3) for j in range(3))
            put(ws, f"{Acol}{rw}",
                f"={Acol}{pv}+{mu_t}+0.5*({quad}+{S2.c(0, 0, q=True)})-{DC.c(0, 0, q=True)}",
                BLK, "0.00000000")
            for j in range(3):
                terms = "+".join(f"{Bc[i]}{pv}*({PHI.c(i, j, q=True)}"
                                 + (f"-{L1.c(i, j, q=True)})" if use_l else ")") for i in range(3))
                put(ws, f"{Bc[j]}{rw}", f"={terms}-{DC.c(1 + j, 0, q=True)}", BLK, "0.00000000")
    ws.freeze_panes = f"B{R0}"

    # ── 결과
    ws = wb.create_sheet("결과")
    put(ws, "A1", "결과 — 실제금리 = 위험중립금리 + 기간프리미엄", TITLE)
    put(ws, "A2", '=IF(VAR!$D$2="BRW","Phi: BRW 편의보정 적용 (권장)",'
                  '"Phi: OLS — 편의보정 없음. 기간프리미엄이 과대추정됩니다.")', RED)
    put(ws, "A3", "모형금리 y(n) = −(A(n) + X(t)'·B(n)) / n × 1200", NOTE)
    head(ws, R0 - 1, ["월", "3년 실제", "3년 모형", "3년 위험중립", "3년 기간프리미엄",
                      "10년 실제", "10년 모형", "10년 위험중립", "10년 기간프리미엄"])
    ws.column_dimensions["A"].width = 10
    for j in range(8):
        ws.column_dimensions[CL(2 + j)].width = 15
    for i in range(T1):
        rw = R0 + i
        put(ws, f"A{rw}", f"=상태변수!A{rw}", GRN).alignment = CEN
        for k, n in enumerate((36, 120)):
            nr = R0 + n
            b = 2 + k * 4
            put(ws, f"{CL(b)}{rw}", f"=무이표금리!{YC[n]}{rw}", GRN, "0.000")
            fit = "+".join(f"상태변수!{XC[j]}{rw}*점화식!${['C', 'D', 'E'][j]}${nr}" for j in range(3))
            put(ws, f"{CL(b + 1)}{rw}", f"=-(점화식!$B${nr}+{fit})/{n}*1200", BLK, "0.000")
            rnf = "+".join(f"상태변수!{XC[j]}{rw}*점화식!${['H', 'I', 'J'][j]}${nr}" for j in range(3))
            put(ws, f"{CL(b + 2)}{rw}", f"=-(점화식!$G${nr}+{rnf})/{n}*1200", BLK, "0.000")
            put(ws, f"{CL(b + 3)}{rw}", f"={CL(b + 1)}{rw}-{CL(b + 2)}{rw}", BLK, "0.000")
    ws.freeze_panes = f"B{R0}"
    for k, lab in enumerate(("3년", "10년")):
        ch = LineChart()
        ch.title = f"국고 {lab}: 실제 vs 위험중립 vs 기간프리미엄 (%)"
        ch.height, ch.width, ch.style = 8.5, 22, 2
        b = 2 + k * 4
        for off in (0, 2, 3):
            ch.add_data(Reference(ws, min_col=b + off, min_row=R0 - 1, max_row=end),
                        titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=R0, max_row=end))
        ws.add_chart(ch, f"K{4 + k * 18}")

    # ── 검증
    ws = wb.create_sheet("검증")
    put(ws, "A1", "검증", TITLE)
    put(ws, "A2", f"기준일 {months[-1]}. 파란 글씨 = 파이썬 계산값. "
                  "VAR 시트 D2 스위치와 같은 Phi 행만 엑셀 값이 채워지고, 나머지 행은 비어 있습니다.", NOTE)
    for j, w in enumerate((9, 8, 18, 12, 12, 12, 46)):
        ws.column_dimensions[CL(1 + j)].width = w
    put(ws, "A4", "① 이 엑셀 vs 같은 모형을 파이썬으로 푼 값 — 차이는 반올림 수준이어야 한다", SUB)
    head(ws, 5, ["Phi", "만기", "항목", "엑셀", "파이썬", "차이(bp)"])
    r = 6
    for tag in ("OLS", "BRW"):
        for n, lab in ((36, "3년"), (120, "10년")):
            b = 2 + (0 if n == 36 else 1) * 4
            for nm, col, key in (("위험중립금리", b + 2, "rn"), ("기간프리미엄", b + 3, "tp")):
                put(ws, f"A{r}", tag, SUB).alignment = CEN
                put(ws, f"B{r}", lab).alignment = CEN
                put(ws, f"C{r}", nm)
                put(ws, f"D{r}", f'=IF(VAR!$D$2="{tag}",결과!{CL(col)}{end},"")', BLK, "0.0000")
                put(ws, f"E{r}", round(ref[tag][lab][key], 6), BLUE, "0.0000")
                put(ws, f"F{r}", f'=IF(D{r}="","",(D{r}-E{r})*100)', BLK, "0.00")
                r += 1
    r += 1
    put(ws, f"A{r}", "② 파이썬 원본(주성분 5팩터·부트스트랩 zero) vs 이 엑셀(관측 3팩터·파금리)", SUB)
    r += 1
    head(ws, r, ["Phi", "만기", "항목", "원본", "이 엑셀", "차이(bp)"])
    r += 1
    orig = {("BRW", "3년"): (3.369, 0.481), ("BRW", "10년"): (3.278, 1.083),
            ("OLS", "3년"): (2.942, 0.908), ("OLS", "10년"): (2.422, 1.939)}
    for tag in ("OLS", "BRW"):
        for lab in ("3년", "10년"):
            for idx, (nm, key) in enumerate((("위험중립금리", "rn"), ("기간프리미엄", "tp"))):
                put(ws, f"A{r}", tag, SUB).alignment = CEN
                put(ws, f"B{r}", lab).alignment = CEN
                put(ws, f"C{r}", nm)
                put(ws, f"D{r}", orig[(tag, lab)][idx], BLUE, "0.000")
                put(ws, f"E{r}", round(ref[tag][lab][key], 4), BLUE, "0.000")
                put(ws, f"F{r}", f"=(E{r}-D{r})*100", BLK, "0.00")
                r += 1
    r += 1
    put(ws, f"A{r}", "③ 모형 적합도 — 관측 금리와 모형 금리의 RMSE (파이썬 계산)", SUB)
    r += 1
    head(ws, r, ["Phi", "만기", "항목", "RMSE(bp)"])
    r += 1
    for tag in ("OLS", "BRW"):
        for lab in ("3년", "10년"):
            put(ws, f"A{r}", tag, SUB).alignment = CEN
            put(ws, f"B{r}", lab).alignment = CEN
            put(ws, f"C{r}", "적합 RMSE")
            put(ws, f"D{r}", round(ref[tag][lab]["rmse_bp"], 2), BLUE, "0.00")
            r += 1
    r += 1
    put(ws, f"A{r}", "④ 기간프리미엄 표본평균 (참고)", SUB)
    r += 1
    head(ws, r, ["Phi", "만기", "항목", "평균(%p)"])
    r += 1
    for tag in ("OLS", "BRW"):
        for lab in ("3년", "10년"):
            put(ws, f"A{r}", tag, SUB).alignment = CEN
            put(ws, f"B{r}", lab).alignment = CEN
            put(ws, f"C{r}", "기간프리미엄 평균")
            put(ws, f"D{r}", round(ref[tag][lab]["tp_mean"], 4), BLUE, "0.000")
            r += 1

    for s in wb.worksheets:
        s.sheet_view.showGridLines = False
    wb.save(out_path)
    return months, ref


if __name__ == "__main__":
    if len(sys.argv) > 1:
        out = Path(sys.argv[1])
    else:
        d = Path.home() / "Desktop" / "Macro"
        out = (d if d.exists() else ROOT) / "ACM_기간프리미엄.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    mo, rf = build(out)
    print(f"[ok] {out}")
    print(f"     표본 {mo[0]} ~ {mo[-1]}")
    for t in ("OLS", "BRW"):
        for lab in ("3년", "10년"):
            d = rf[t][lab]
            print(f"     {t} {lab}: 위험중립 {d['rn']:.3f}%  TP {d['tp']:+.3f}%p  적합 {d['rmse_bp']:.2f}bp")
