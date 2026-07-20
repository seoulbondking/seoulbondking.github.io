"""CPI 품목 단위 기여도 검증 엑셀 생성.

품목별 기여도의 합이 Headline / Core / 생활물가지수의 MoM·YoY와 일치하는지
수식으로 점검하는 워크북을 만든다.

사전 준비 (품목별 지수 필요):
    python fetch.py --full kr_cpi_items
실행:
    python tools/build_cpi_check.py
결과:
    C:/Users/infomax/Desktop/Macro/CPI기여도_검증.xlsx
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import os
ROOT = Path(__file__).parent.parent
DATA = ROOT / "docs" / "data"
OUT = Path(os.environ.get("CPI_CHECK_OUT", "C:/Users/infomax/Desktop/Macro/CPI기여도_검증.xlsx"))
N_MONTHS = 25          # 최근 25개월 (전부 2022 가중치 + 리레이팅 구간)
LINK = "2021-12-31"    # 이후 구간: 2022 가중치 + 2022년 연평균=100 리레이팅 (KR_Macro 방식)
import calendar as _cal
MON22 = [f"2022-{m:02d}-{_cal.monthrange(2022, m)[1]:02d}" for m in range(1, 13)]


def load(fid):
    raw = (DATA / f"{fid}.json").read_text(encoding="utf-8")
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        d, _ = json.JSONDecoder().raw_decode(raw)
        return d


def smap(ind, name):
    for s in ind["series"]:
        if s["name"] == name:
            return {p["d"]: p["v"] for p in s["data"]}
    for s in ind["series"]:
        if name in s["name"]:
            return {p["d"]: p["v"] for p in s["data"]}
    sys.exit(f"시리즈 없음: {name}")


def main():
    struct = json.loads(re.sub(r"^window\.__CPI_STRUCT__=|;$", "",
                               (DATA / "cpi_struct.js").read_text(encoding="utf-8")))
    cpi = load("kr_cpi"); core = load("kr_cpi_core"); liv = load("kr_cpi_living")
    try:
        items = load("kr_cpi_items")
    except FileNotFoundError:
        sys.exit("kr_cpi_items.json 이 없습니다. 먼저: python fetch.py --full kr_cpi_items")

    base = lambda n: re.sub(r" \[[^\]]*\]$", "", n)
    imap = {base(s["name"]): {p["d"]: p["v"] for p in s["data"]} for s in items["series"]}

    H = smap(cpi, "총지수")
    C = smap(core, "식료품 및 에너지제외 지수")
    L = smap(liv, "생활물가지수")
    dates = sorted(H)[-N_MONTHS:]
    leafs = [nd for nd in struct if not nd["agg"]]
    missing = [nd["n"] for nd in leafs if nd["n"] not in imap]
    if missing:
        print(f"주의: 품목 지수 미확보 {len(missing)}개 (예: {missing[:5]}) — 해당 품목은 빈칸")

    wb = openpyxl.Workbook()
    F = lambda **k: Font(name="맑은 고딕", **k)
    HD = PatternFill("solid", fgColor="1A2B4A")

    def hstyle(c):
        c.font = F(bold=True, color="FFFFFF"); c.fill = HD
        c.alignment = Alignment(horizontal="center")

    # ── 설명 ──────────────────────────────────────────────
    ws = wb.active; ws.title = "설명"
    for i, (a, b) in enumerate([
        ("CPI 품목 단위 기여도 검증", ""), ("", ""),
        ("원리", "각 지수 = Σ(가중치×품목지수)/가중치합 이므로, 품목 기여도의 합 = 지수 변화율"),
        ("기여도 공식", "기여도_i = (P_i,t − P_i,t−k) × f_i / P_지수,t−k × w_i/W × 100   (k=1 MoM, 12 YoY)"),
        ("리레이팅 f_i", "2022년 연평균=100 재기준 계수 = 지수총(2022평균) / 품목지수(2022평균) — KR_Macro CPI(상세) 방식"),
        ("구성", "Headline 458품목(W=1000) · Core 309품목(W=782.2) · 생활물가 144품목(W=528.4)"),
        ("Core 품목", "전체 − 식료품·비주류음료(140) − 에너지 9품목(휘발유·경유·자동차용LPG·등유·부탄가스·전기료·도시가스·취사용LPG·지역난방비)"),
        ("가중치", f"2022년 기준 + 리레이팅 (기간 {dates[0][:7]}~{dates[-1][:7]})"),
        ("점검", "각 시트 상단 '차이' 행이 ±0.05%p 이내면 정상 (공표 지수 반올림 한계)"),
        ("생성일", datetime.now().strftime("%Y-%m-%d")),
    ], 1):
        ws.cell(row=i, column=1, value=a).font = F(bold=(i == 1), size=14 if i == 1 else 11)
        ws.cell(row=i, column=2, value=b).font = F()
    ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 110

    # ── 지수 시트 (원자료) ────────────────────────────────
    zs = wb.create_sheet("지수")
    hstyle(zs.cell(row=1, column=1, value="구분"))
    for j, d in enumerate(dates):
        c = zs.cell(row=1, column=2 + j, value=datetime.strptime(d, "%Y-%m-%d"))
        hstyle(c); c.number_format = "yyyy-mm"
    for r, (nm, mm) in enumerate([("Headline총지수", H), ("Core지수", C), ("생활물가지수", L)], 2):
        zs.cell(row=r, column=1, value=nm).font = F(bold=True)
        for j, d in enumerate(dates):
            zs.cell(row=r, column=2 + j, value=mm.get(d)).number_format = "0.00"
    item_row = {}
    for k, nd in enumerate(leafs):
        r = 6 + k
        item_row[nd["n"]] = r
        zs.cell(row=r, column=1, value=nd["n"]).font = F()
        mm = imap.get(nd["n"], {})
        for j, d in enumerate(dates):
            zs.cell(row=r, column=2 + j, value=mm.get(d)).number_format = "0.00"
    zs.freeze_panes = "B2"
    zs.column_dimensions["A"].width = 20

    # ── 기여도 시트 생성 ──────────────────────────────────
    def avg22(mm):
        vals = [mm.get(d) for d in MON22]
        return sum(vals) / 12 if all(v is not None for v in vals) else None

    def rebase_factor(nd, tot_map):
        """f_i = 지수총(2022평균)/품목지수(2022평균). 신규 품목은 첫 관측월로 연결."""
        mm = imap.get(nd["n"])
        TA = avg22(tot_map)
        if not mm or TA is None:
            return None
        pa = avg22(mm)
        if pa:
            return TA / pa
        for d0 in sorted(mm):
            if d0 > LINK and mm.get(d0) and tot_map.get(d0):
                return tot_map[d0] / mm[d0]
        return None

    def contrib_sheet(name, tot_row, tot_map, memb_key, W, lag):
        ws = wb.create_sheet(name)
        members = [nd for nd in leafs if memb_key(nd)]
        ws.cell(row=1, column=1, value=name).font = F(bold=True, size=12)
        ws.cell(row=2, column=1, value="가중치합 W").font = F(bold=True)
        ws.cell(row=2, column=2, value=W).number_format = "0.0"
        hstyle(ws.cell(row=4, column=1, value="항목"))
        hstyle(ws.cell(row=4, column=2, value="가중치22"))
        hstyle(ws.cell(row=4, column=3, value="리레이팅f"))
        for j, d in enumerate(dates):
            c = ws.cell(row=4, column=4 + j, value=datetime.strptime(d, "%Y-%m-%d"))
            hstyle(c); c.number_format = "yyyy-mm"
        n0 = 8                      # 품목 시작 행
        nL = n0 + len(members) - 1
        for lbl, r in [("실제 변화율 (%)", 5), ("품목 기여도 합 (%p)", 6), ("차이", 7)]:
            ws.cell(row=r, column=1, value=lbl).font = F(bold=True)
        for j in range(len(dates)):
            if j < lag:
                continue
            col = 4 + j
            cl = get_column_letter(col)
            zcl = get_column_letter(2 + j)          # 지수 시트의 같은 날짜 열
            zpl = get_column_letter(2 + j - lag)    # 지수 시트의 lag 이전 열
            ws.cell(row=5, column=col,
                    value=f"=(지수!{zcl}{tot_row}/지수!{zpl}{tot_row}-1)*100").number_format = "0.00"
            ws.cell(row=6, column=col,
                    value=f"=SUM({cl}{n0}:{cl}{nL})").number_format = "0.00"
            ws.cell(row=7, column=col, value=f"={cl}5-{cl}6").number_format = "0.00"
        for k, nd in enumerate(members):
            r = n0 + k
            zr = item_row[nd["n"]]
            ws.cell(row=r, column=1, value=nd["n"]).font = F()
            ws.cell(row=r, column=2, value=nd["w22"]).number_format = "0.0"
            ws.cell(row=r, column=3, value=rebase_factor(nd, tot_map)).number_format = "0.0000"
            for j in range(len(dates)):
                if j < lag:
                    continue
                col = 4 + j
                zcl = get_column_letter(2 + j)
                zpl = get_column_letter(2 + j - lag)
                ws.cell(row=r, column=col,
                        value=f"=IFERROR((지수!{zcl}{zr}-지수!{zpl}{zr})*$C{r}/지수!{zpl}${tot_row}*$B{r}/$B$2*100,\"\")"
                        ).number_format = "0.00"
        ws.freeze_panes = "D5"
        ws.column_dimensions["A"].width = 20
        return members

    W_H = sum(nd["w22"] or 0 for nd in leafs)
    W_C = sum(nd["w22"] or 0 for nd in leafs if nd["core"])
    W_L = sum(nd["w22"] or 0 for nd in leafs if nd["liv"])
    contrib_sheet("Headline_MoM", 2, H, lambda nd: True, W_H, 1)
    contrib_sheet("Headline_YoY", 2, H, lambda nd: True, W_H, 12)
    contrib_sheet("Core_MoM", 3, C, lambda nd: nd["core"], W_C, 1)
    contrib_sheet("Core_YoY", 3, C, lambda nd: nd["core"], W_C, 12)
    contrib_sheet("생활물가_MoM", 4, L, lambda nd: nd["liv"], W_L, 1)
    contrib_sheet("생활물가_YoY", 4, L, lambda nd: nd["liv"], W_L, 12)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"저장: {OUT}")

    # ── 파이썬 병렬 검증 (엑셀 열기 전 사전 확인) ─────────
    def check(tot, memb_key, W, lag, label):
        diffs = []
        for j in range(lag, len(dates)):
            d, p = dates[j], dates[j - lag]
            if tot.get(d) is None or tot.get(p) is None:
                continue
            actual = (tot[d] / tot[p] - 1) * 100
            ssum = 0
            for nd in leafs:
                if not memb_key(nd):
                    continue
                mm = imap.get(nd["n"])
                if not mm or mm.get(d) is None or mm.get(p) is None:
                    continue
                f = rebase_factor(nd, tot)
                if f is None:
                    continue
                ssum += (mm[d] - mm[p]) * f / tot[p] * (nd["w22"] or 0) / W * 100
            diffs.append(abs(actual - ssum))
        print(f"  {label}: |차이| 최대 {max(diffs):.4f} · 평균 {sum(diffs)/len(diffs):.4f} %p")

    print("검증 (파이썬 계산):")
    check(H, lambda nd: True, W_H, 1, "Headline MoM")
    check(H, lambda nd: True, W_H, 12, "Headline YoY")
    check(C, lambda nd: nd["core"], W_C, 1, "Core MoM")
    check(C, lambda nd: nd["core"], W_C, 12, "Core YoY")
    check(L, lambda nd: nd["liv"], W_L, 1, "생활물가 MoM")
    check(L, lambda nd: nd["liv"], W_L, 12, "생활물가 YoY")


if __name__ == "__main__":
    main()
