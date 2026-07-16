"""부동산원 R-ONE 주간 아파트 3개 통계표의 CLS_ID ↔ 지역명 매핑을 엑셀로 저장.

사용법 (PC의 venv 파이썬):
    python tools/reb_export_cls.py

결과: tools/reb_cls_codes.xlsx
  - 시트 '통합': 지역명 × 3개 통계표 CLS_ID 를 한눈에 (대시보드 카테고리 포함)
  - 시트별(매매/전세/수급): 각 표의 CLS_ID · 지역명 · 항목(ITM) 전체
"""
import os
import sys
import xml.etree.ElementTree as ET
from datetime import date
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent
BASE_URL = "https://www.reb.or.kr/r-one/openapi/SttsApiTblData.do"

TABLES = [
    ("매매", "T244183132827305"),
    ("전세", "T247713133046872"),
    ("수급", "T248163133074619"),
]

# 대시보드와 동일한 카테고리 분류 (index.html 과 맞춰둘 것)
GYEONGGI = set("수원시 성남시 고양시 용인시 부천시 안산시 안양시 남양주시 화성시 평택시 의정부시 시흥시 파주시 김포시 광명시 광주시 군포시 오산시 이천시 양주시 안성시 구리시 포천시 의왕시 하남시 여주시 동두천시 과천시 양평군 가평군 연천군 장안구 권선구 팔달구 영통구 수정구 중원구 분당구 만안구 동안구 상록구 단원구 덕양구 일산동구 일산서구 처인구 기흥구 수지구 원미구 소사구 오정구 만세구 효행구 병점구 동탄구 경부1권 경부2권 서해안권 동부1권 동부2권 경의권 경원권".split())
INCHEON = set("남동구 부평구 계양구 연수구 미추홀구 강화군 옹진구".split())
SEOUL_GU = set("종로구 용산구 성동구 광진구 동대문구 중랑구 성북구 강북구 도봉구 노원구 은평구 서대문구 마포구 양천구 구로구 금천구 영등포구 동작구 관악구 서초구 강남구 송파구 강동구".split())


# 서울의 겹치는 자치구 CLS_ID (매매 통계표 기준) → 서울 자치구로 강제 분류
SEOUL_GU_CID = {"50044": "중구", "50061": "강서구"}


def category(n: str, cid: str = "") -> str:
    if cid in SEOUL_GU_CID:
        return "서울 자치구"
    if n in ("전국", "수도권", "지방", "지방권"):
        return "전국·수도권·지방"
    if n in ("5대광역시", "6대광역시", "8개도", "9개도"):
        return "전국 집계권역"
    if n == "서울":
        return "서울"
    if n in ("강북지역", "강남지역"):
        return "서울 강북·강남지역"
    if n in ("도심권", "동북권", "서북권", "서남권", "동남권"):
        return "서울 5대권역"
    if n in SEOUL_GU:
        return "서울 자치구"
    if n == "인천" or n in INCHEON:
        return "인천"
    if n == "경기" or n in GYEONGGI:
        return "경기"
    if n in ("세종", "세종시"):
        return "세종"
    return "지방 (광역시·도·시군구)"


CAT_ORDER = ["전국·수도권·지방", "전국 집계권역", "서울", "서울 강북·강남지역",
             "서울 5대권역", "서울 자치구", "인천", "경기", "세종", "지방 (광역시·도·시군구)"]


def load_dotenv():
    env = ROOT / ".env"
    if env.exists():
        for line in env.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip())


def fetch_cls(key: str, statbl_id: str) -> dict:
    """{CLS_ID: (지역명, set(ITM))} 반환. 올해 데이터만 조회(코드 나열 목적)."""
    start = f"{date.today().year}01"
    found = {}
    for page in range(1, 21):
        params = {"KEY": key, "pIndex": str(page), "pSize": "1000",
                  "STATBL_ID": statbl_id, "DTACYCLE_CD": "WK", "START_WRTTIME": start}
        resp = requests.get(BASE_URL, params=params, timeout=60)
        root = ET.fromstring(resp.content)
        rows = root.findall(".//row")
        if not rows:
            if page == 1:
                msg = (root.findtext(".//message") or root.findtext(".//MESSAGE") or "").strip()
                print(f"  [{statbl_id}] 데이터 없음 {msg}")
            break
        for r in rows:
            cid = (r.findtext("CLS_ID") or "").strip()
            nm = (r.findtext("CLS_NM") or "").strip()
            itm = (r.findtext("ITM_NM") or "").strip()
            if cid:
                e = found.setdefault(cid, [nm, set()])
                if itm:
                    e[1].add(itm)
        if len(rows) < 1000:
            break
    return found


def main():
    load_dotenv()
    key = os.environ.get("REB_API_KEY", "").strip()
    if not key:
        sys.exit("REB_API_KEY 가 없습니다 (.env 확인)")

    per_table = {}          # 표라벨 → {cid: (nm, itms)}
    for label, sid in TABLES:
        print(f"조회 중: {label} ({sid})")
        per_table[label] = fetch_cls(key, sid)
        print(f"   → {len(per_table[label])}개 지역")

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A2B4A")

    # ── 통합 시트: 지역명 기준, 표별 CLS_ID 나란히 ──────────────
    ws = wb.active
    ws.title = "통합"
    ws.append(["카테고리", "지역명", "매매 CLS_ID", "전세 CLS_ID", "수급 CLS_ID", "항목(ITM)"])
    # 지역명 → {표: cid}
    by_name = {}
    itm_by_name = {}
    for label, _ in TABLES:
        for cid, (nm, itms) in per_table[label].items():
            by_name.setdefault(nm, {})[label] = cid
            itm_by_name.setdefault(nm, set()).update(itms)
    def cat_of(nm, cids):   # 매매 CLS_ID 기준으로 서울 자치구 보정
        return category(nm, cids.get("매매", ""))
    rows = sorted(by_name.items(),
                  key=lambda kv: (CAT_ORDER.index(cat_of(kv[0], kv[1])), kv[0]))
    for nm, cids in rows:
        ws.append([cat_of(nm, cids), nm, cids.get("매매", ""), cids.get("전세", ""),
                   cids.get("수급", ""), " / ".join(sorted(itm_by_name.get(nm, [])))])

    # ── 표별 상세 시트 ─────────────────────────────────────────
    for label, _ in TABLES:
        s = wb.create_sheet(label)
        s.append(["CLS_ID", "지역명", "카테고리", "항목(ITM)"])
        for cid in sorted(per_table[label]):
            nm, itms = per_table[label][cid]
            s.append([cid, nm, category(nm, cid), " / ".join(sorted(itms))])

    # 서식
    for s in wb.worksheets:
        for c in s[1]:
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        s.freeze_panes = "A2"
        for col in s.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=8)
            s.column_dimensions[col[0].column_letter].width = min(max(w * 1.6, 10), 40)

    out = ROOT / "tools" / "reb_cls_codes.xlsx"
    wb.save(out)
    print(f"\n저장 완료: {out}")


if __name__ == "__main__":
    main()
